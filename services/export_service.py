# services/export_service.py
import pandas as pd
import streamlit as st
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
import json
from io import BytesIO
import base64

# PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    st.warning("לייצוא PDF חסרה הספריה reportlab. התקן עם: pip install reportlab")

# Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    st.warning("לייצוא Excel חסרה הספריה openpyxl. התקן עם: pip install openpyxl")

from config import Config

class ExportService:
    def __init__(self):
        self.pdf_available = PDF_AVAILABLE
        self.excel_available = EXCEL_AVAILABLE
    
    def export_to_excel_advanced(self, df: pd.DataFrame, filename: Optional[str] = None) -> BytesIO:
        """ייצוא מתקדם לאקסל עם עיצוב וגרפים"""
        if not self.excel_available:
            raise Exception("חסרות ספריות Excel")
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"giraffe_quality_report_{timestamp}.xlsx"
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # גיליון נתונים גולמיים
            self._write_raw_data_sheet(df, writer)
            
            # גיליון סיכום
            self._write_summary_sheet(df, writer)
            
            # גיליון ניתוח סניפים
            self._write_branch_analysis_sheet(df, writer)
            
            # גיליון ניתוח מנות
            self._write_dish_analysis_sheet(df, writer)
            
            # גיליון מגמות
            self._write_trends_sheet(df, writer)
        
        output.seek(0)
        return output
    
    def _write_raw_data_sheet(self, df: pd.DataFrame, writer):
        """כתיבת גיליון נתונים גולמיים"""
        # הכנת הנתונים
        export_df = df[[
            'created_at', 'branch', 'chef_name', 'dish_name', 'overall_score',
            'taste_score', 'appearance_score', 'temperature_score',
            'preparation_time_score', 'portion_size_score', 'notes'
        ]].copy()
        
        # תרגום כותרות
        export_df = export_df.rename(columns={
            'created_at': 'תאריך ושעה',
            'branch': 'סניף',
            'chef_name': 'שם טבח',
            'dish_name': 'שם מנה',
            'overall_score': 'ציון כללי',
            'taste_score': 'ציון טעם',
            'appearance_score': 'ציון מראה',
            'temperature_score': 'ציון טמפרטורה',
            'preparation_time_score': 'ציון זמן הכנה',
            'portion_size_score': 'ציון כמות',
            'notes': 'הערות'
        })
        
        # המרת תאריכים
        export_df['תאריך ושעה'] = export_df['תאריך ושעה'].dt.strftime('%d/%m/%Y %H:%M')
        
        # כתיבה לגיליון
        export_df.to_excel(writer, sheet_name='נתונים גולמיים', index=False)
        
        # עיצוב
        workbook = writer.book
        worksheet = writer.sheets['נתונים גולמיים']
        
        # עיצוב כותרות
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # הדגשת ציונים נמוכים
        red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
        
        score_col = 5  # עמודת הציון הכללי
        for row in range(2, len(export_df) + 2):
            cell = worksheet.cell(row=row, column=score_col)
            if cell.value and isinstance(cell.value, (int, float)):
                if cell.value <= 6:
                    cell.fill = red_fill
                elif cell.value <= 7:
                    cell.fill = yellow_fill
        
        # התאמת רוחב עמודות
        column_widths = [20, 15, 20, 25, 12, 12, 12, 12, 12, 12, 40]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[worksheet.cell(1, i).column_letter].width = width
    
    def _write_summary_sheet(self, df: pd.DataFrame, writer):
        """כתיבת גיליון סיכום"""
        summary_data = []
        
        # כותרת
        summary_data.append(['דוח סיכום איכות ג\'ירף'])
        summary_data.append([f'נוצר בתאריך: {datetime.now().strftime("%d/%m/%Y %H:%M")}'])
        summary_data.append([''])
        
        # סטטיסטיקות כלליות
        total_checks = len(df)
        avg_score = df['overall_score'].mean()
        unique_branches = df['branch'].nunique()
        unique_dishes = df['dish_name'].nunique()
        
        summary_data.extend([
            ['📊 סטטיסטיקות כלליות'],
            ['סה״כ בדיקות איכות:', total_checks],
            ['ממוצע ציונים:', f'{avg_score:.2f}'],
            ['מספר סניפים:', unique_branches],
            ['מספר מנות שנבדקו:', unique_dishes],
            ['']
        ])
        
        # ביצועי סניפים
        branch_stats = df.groupby('branch')['overall_score'].agg(['mean', 'count', 'min', 'max']).round(2)
        branch_stats = branch_stats.sort_values('mean', ascending=False)
        
        summary_data.append(['🏪 ביצועי סניפים'])
        summary_data.append(['סניף', 'ממוצע', 'בדיקות', 'מינימום', 'מקסימום'])
        
        for branch, row in branch_stats.iterrows():
            summary_data.append([
                branch, 
                row['mean'], 
                int(row['count']), 
                row['min'], 
                row['max']
            ])
        
        summary_data.append([''])
        
        # מנות בעייתיות
        low_score_dishes = df[df['overall_score'] <= 6].groupby('dish_name')['overall_score'].agg(['mean', 'count']).round(2)
        if not low_score_dishes.empty:
            summary_data.append(['🚨 מנות הדורשות תשומת לב (ממוצע 6 ומטה)'])
            summary_data.append(['מנה', 'ממוצע', 'מספר בדיקות'])
            
            for dish, row in low_score_dishes.iterrows():
                summary_data.append([dish, row['mean'], int(row['count'])])
        
        # כתיבה לגיליון
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='סיכום', index=False, header=False)
        
        # עיצוב
        worksheet = writer.sheets['סיכום']
        
        # כותרת ראשית
        worksheet['A1'].font = Font(size=16, bold=True, color="2E8B57")
        worksheet['A1'].alignment = Alignment(horizontal="center")
        
        # כותרות משנה
        for row in [4, 11]:
            if row <= len(summary_data):
                cell = worksheet[f'A{row}']
                cell.font = Font(size=12, bold=True, color="2E8B57")
    
    def _write_branch_analysis_sheet(self, df: pd.DataFrame, writer):
        """ניתוח מפורט לפי סניפים"""
        analysis_data = []
        
        # כותרת
        analysis_data.append(['ניתוח מפורט לפי סניפים'])
        analysis_data.append([''])
        
        for branch in Config.BRANCHES:
            branch_df = df[df['branch'] == branch]
            if branch_df.empty:
                continue
            
            # נתונים בסיסיים
            analysis_data.extend([
                [f'🏪 {branch}'],
                ['סה״כ בדיקות:', len(branch_df)],
                ['ממוצע ציונים:', f'{branch_df["overall_score"].mean():.2f}'],
                ['ציון מינימלי:', branch_df["overall_score"].min()],
                ['ציון מקסימלי:', branch_df["overall_score"].max()],
                ['']
            ])
            
            # טבחים מובילים
            chef_stats = branch_df.groupby('chef_name')['overall_score'].agg(['mean', 'count']).round(2)
            chef_stats = chef_stats[chef_stats['count'] >= 2].sort_values('mean', ascending=False)
            
            if not chef_stats.empty:
                analysis_data.append(['👨‍🍳 ביצועי טבחים:'])
                analysis_data.append(['טבח', 'ממוצע', 'בדיקות'])
                
                for chef, row in chef_stats.head(5).iterrows():
                    analysis_data.append([chef, row['mean'], int(row['count'])])
                
                analysis_data.append([''])
            
            # מנות פופולריות
            dish_stats = branch_df.groupby('dish_name')['overall_score'].agg(['mean', 'count']).round(2)
            dish_stats = dish_stats.sort_values('count', ascending=False)
            
            if not dish_stats.empty:
                analysis_data.append(['🍽️ מנות פופולריות:'])
                analysis_data.append(['מנה', 'ממוצע', 'בדיקות'])
                
                for dish, row in dish_stats.head(5).iterrows():
                    analysis_data.append([dish, row['mean'], int(row['count'])])
                
                analysis_data.append([''])
            
            analysis_data.append(['─' * 50])
            analysis_data.append([''])
        
        # כתיבה לגיליון
        analysis_df = pd.DataFrame(analysis_data)
        analysis_df.to_excel(writer, sheet_name='ניתוח סניפים', index=False, header=False)
        
        # עיצוב בסיסי
        worksheet = writer.sheets['ניתוח סניפים']
        worksheet['A1'].font = Font(size=14, bold=True, color="2E8B57")
    
    def _write_dish_analysis_sheet(self, df: pd.DataFrame, writer):
        """ניתוח מפורט לפי מנות"""
        dish_stats = df.groupby('dish_name').agg({
            'overall_score': ['mean', 'count', 'min', 'max', 'std'],
            'branch': lambda x: ', '.join(x.unique())
        }).round(2)
        
        dish_stats.columns = ['ממוצע', 'בדיקות', 'מינימום', 'מקסימום', 'סטיית תקן', 'סניפים']
        dish_stats = dish_stats.sort_values('ממוצע', ascending=False)
        
        # הוספת דירוג
        dish_stats['דירוג'] = range(1, len(dish_stats) + 1)
        dish_stats = dish_stats[['דירוג', 'ממוצע', 'בדיקות', 'מינימום', 'מקסימום', 'סטיית תקן', 'סניפים']]
        
        # כתיבה לגיליון
        dish_stats.to_excel(writer, sheet_name='ניתוח מנות')
        
        # עיצוב
        worksheet = writer.sheets['ניתוח מנות']
        
        # כותרות
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
    
    def _write_trends_sheet(self, df: pd.DataFrame, writer):
        """גיליון מגמות"""
        if len(df) < 10:  # לא מספיק נתונים למגמות
            trend_data = [['אין מספיק נתונים לניתוח מגמות']]
        else:
            # ניתוח מגמות יומיות
            df_daily = df.copy()
            df_daily['date'] = df_daily['created_at'].dt.date
            daily_stats = df_daily.groupby('date')['overall_score'].agg(['mean', 'count']).round(2)
            
            trend_data = [
                ['ניתוח מגמות יומיות'],
                [''],
                ['תאריך', 'ממוצע יומי', 'מספר בדיקות']
            ]
            
            for date, row in daily_stats.iterrows():
                trend_data.append([
                    date.strftime('%d/%m/%Y'),
                    row['mean'],
                    int(row['count'])
                ])
        
        # כתיבה לגיליון
        trend_df = pd.DataFrame(trend_data)
        trend_df.to_excel(writer, sheet_name='מגמות', index=False, header=False)
    
    def export_to_pdf(self, df: pd.DataFrame, filename: Optional[str] = None) -> BytesIO:
        """ייצוא לPDF"""
        if not self.pdf_available:
            raise Exception("חסרות ספריות PDF")
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"giraffe_quality_report_{timestamp}.pdf"
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=18)
        
        # הגדרת סטיילים
        styles = getSampleStyleSheet()
        
        # סטייל לכותרת
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2E8B57')
        )
        
        # סטייל לטקסט רגיל
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_RIGHT,
            wordWrap='RTL'
        )
        
        # בניית התוכן
        story = []
        
        # כותרת
        story.append(Paragraph("🦒 דוח איכות מזון - רשת ג'ירף", title_style))
        story.append(Paragraph(f"נוצר בתאריך: {datetime.now().strftime('%d/%m/%Y %H:%M')}", normal_style))
        story.append(Spacer(1, 20))
        
        # סיכום כללי
        total_checks = len(df)
        avg_score = df['overall_score'].mean()
        
        summary_text = f"""
        <b>סיכום כללי:</b><br/>
        • סה"כ בדיקות איכות: {total_checks}<br/>
        • ממוצע ציונים: {avg_score:.2f}<br/>
        • מספר סניפים: {df['branch'].nunique()}<br/>
        • מספר מנות שנבדקו: {df['dish_name'].nunique()}<br/>
        """
        
        story.append(Paragraph(summary_text, normal_style))
        story.append(Spacer(1, 20))
        
        # טבלת ביצועי סניפים
        branch_stats = df.groupby('branch')['overall_score'].agg(['mean', 'count']).round(2)
        branch_stats = branch_stats.sort_values('mean', ascending=False)
        
        story.append(Paragraph("<b>ביצועי סניפים:</b>", normal_style))
        story.append(Spacer(1, 10))
        
        # יצירת טבלה
        table_data = [['סניף', 'ממוצע', 'בדיקות']]
        for branch, row in branch_stats.iterrows():
            table_data.append([
                branch,
                f'{row["mean"]:.2f}',
                str(int(row['count']))
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E8B57')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
        
        # מנות בעייתיות
        low_score_dishes = df[df['overall_score'] <= 6]
        if not low_score_dishes.empty:
            problem_dishes = low_score_dishes.groupby('dish_name')['overall_score'].agg(['mean', 'count']).round(2)
            
            story.append(Paragraph("<b>מנות הדורשות תשומת לב (ציון 6 ומטה):</b>", normal_style))
            story.append(Spacer(1, 10))
            
            problem_table_data = [['מנה', 'ממוצע', 'בדיקות']]
            for dish, row in problem_dishes.iterrows():
                problem_table_data.append([
                    dish,
                    f'{row["mean"]:.2f}',
                    str(int(row['count']))
                ])
            
            problem_table = Table(problem_table_data)
            problem_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.red),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.pink),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(problem_table)
        
        # יצירת PDF
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def create_weekly_report_pdf(self, df: pd.DataFrame, week_start: datetime) -> BytesIO:
        """יצירת דוח שבועי PDF"""
        if not self.pdf_available:
            raise Exception("חסרות ספריות PDF")
        
        # סינון נתונים לשבוע
        week_end = week_start + timedelta(days=7)
        weekly_df = df[
            (df['created_at'] >= week_start) & 
            (df['created_at'] < week_end)
        ]
        
        if weekly_df.empty:
            raise Exception("אין נתונים לשבוע הנבחר")
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'WeeklyTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2E8B57')
        )
        
        story = []
        
        # כותרת
        story.append(Paragraph(
            f"📋 דוח שבועי - {week_start.strftime('%d/%m/%Y')} עד {week_end.strftime('%d/%m/%Y')}", 
            title_style
        ))
        
        # סיכום שבועי
        total_weekly = len(weekly_df)
        avg_weekly = weekly_df['overall_score'].mean()
        active_branches = weekly_df['branch'].nunique()
        
        summary = f"""
        <b>סיכום השבוע:</b><br/>
        • סה"כ בדיקות: {total_weekly}<br/>
        • ממוצע ציונים: {avg_weekly:.2f}<br/>
        • סניפים פעילים: {active_branches}<br/>
        """
        
        story.append(Paragraph(summary, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # ביצועי סניפים השבוע
        weekly_branches = weekly_df.groupby('branch')['overall_score'].agg(['mean', 'count']).round(2)
        weekly_branches = weekly_branches.sort_values('mean', ascending=False)
        
        story.append(Paragraph("<b>ביצועי סניפים השבוע:</b>", styles['Normal']))
        
        branch_table_data = [['דירוג', 'סניף', 'ממוצע', 'בדיקות']]
        for i, (branch, row) in enumerate(weekly_branches.iterrows(), 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else ""
            branch_table_data.append([
                f'{i} {emoji}',
                branch,
                f'{row["mean"]:.2f}',
                str(int(row['count']))
            ])
        
        branch_table = Table(branch_table_data)
        branch_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E8B57')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(branch_table)
        story.append(PageBreak())
        
        # פירוט לפי יום
        story.append(Paragraph("<b>פירוט יומי:</b>", styles['Heading2']))
        
        daily_stats = weekly_df.copy()
        daily_stats['date'] = daily_stats['created_at'].dt.date
        daily_summary = daily_stats.groupby('date').agg({
            'overall_score': ['mean', 'count'],
            'branch': lambda x: len(x.unique())
        }).round(2)
        
        daily_table_data = [['תאריך', 'ממוצע', 'בדיקות', 'סניפים']]
        for date, row in daily_summary.iterrows():
            daily_table_data.append([
                date.strftime('%d/%m/%Y'),
                f'{row[("overall_score", "mean")]:.2f}',
                str(int(row[("overall_score", "count")])),
                str(int(row[("branch", "<lambda>")]))
            ])
        
        daily_table = Table(daily_table_data)
        daily_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E8B57')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(daily_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer
    
    def export_to_csv_enhanced(self, df: pd.DataFrame, include_metadata: bool = True) -> str:
        """ייצוא CSV משופר"""
        export_df = df.copy()
        
        # הוספת מטאדאטה
        if include_metadata:
            export_df['export_date'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            export_df['week_number'] = export_df['created_at'].dt.isocalendar().week
            export_df['day_of_week'] = export_df['created_at'].dt.day_name()
        
        # עיבוד נתונים לייצוא
        export_df['created_at'] = export_df['created_at'].dt.strftime('%d/%m/%Y %H:%M')
        
        # סדר עמודות מועדף
        preferred_columns = [
            'created_at', 'branch', 'chef_name', 'dish_name', 'overall_score',
            'taste_score', 'appearance_score', 'temperature_score',
            'preparation_time_score', 'portion_size_score', 'notes'
        ]
        
        # הוספת עמודות נוספות אם קיימות
        if include_metadata:
            preferred_columns.extend(['export_date', 'week_number', 'day_of_week'])
        
        # בחירת עמודות קיימות בלבד
        available_columns = [col for col in preferred_columns if col in export_df.columns]
        export_df = export_df[available_columns]
        
        return export_df.to_csv(index=False, encoding='utf-8-sig')
    
    def create_summary_json(self, df: pd.DataFrame) -> str:
        """יצירת סיכום JSON מפורט"""
        summary = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'total_records': len(df),
                'date_range': {
                    'start': df['created_at'].min().isoformat() if not df.empty else None,
                    'end': df['created_at'].max().isoformat() if not df.empty else None
                }
            },
            'overall_statistics': {
                'total_checks': len(df),
                'average_score': float(df['overall_score'].mean()) if not df.empty else 0,
                'score_distribution': df['overall_score'].value_counts().to_dict() if not df.empty else {},
                'unique_branches': int(df['branch'].nunique()) if not df.empty else 0,
                'unique_dishes': int(df['dish_name'].nunique()) if not df.empty else 0,
                'unique_chefs': int(df['chef_name'].nunique()) if not df.empty else 0
            }
        }
        
        if not df.empty:
            # סטטיסטיקות לפי סניף
            branch_stats = df.groupby('branch')['overall_score'].agg([
                'mean', 'count', 'min', 'max', 'std'
            ]).round(2)
            
            summary['branch_statistics'] = {}
            for branch, stats in branch_stats.iterrows():
                summary['branch_statistics'][branch] = {
                    'average_score': float(stats['mean']),
                    'total_checks': int(stats['count']),
                    'min_score': int(stats['min']),
                    'max_score': int(stats['max']),
                    'standard_deviation': float(stats['std']) if pd.notna(stats['std']) else 0
                }
            
            # סטטיסטיקות לפי מנה
            dish_stats = df.groupby('dish_name')['overall_score'].agg([
                'mean', 'count'
            ]).round(2)
            
            summary['dish_statistics'] = {}
            for dish, stats in dish_stats.iterrows():
                summary['dish_statistics'][dish] = {
                    'average_score': float(stats['mean']),
                    'total_checks': int(stats['count'])
                }
            
            # בדיקות אחרונות (7 ימים)
            recent_df = df[df['created_at'] >= (datetime.now() - timedelta(days=7))]
            if not recent_df.empty:
                summary['recent_activity'] = {
                    'last_7_days': {
                        'total_checks': len(recent_df),
                        'average_score': float(recent_df['overall_score'].mean()),
                        'active_branches': int(recent_df['branch'].nunique())
                    }
                }
        
        return json.dumps(summary, ensure_ascii=False, indent=2)
    
    def get_export_filename(self, file_type: str, prefix: str = "giraffe_quality") -> str:
        """יצירת שם קובץ לייצוא"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{timestamp}.{file_type}"

# יצירת אינסטנס גלובלי
export_service = ExportService()
